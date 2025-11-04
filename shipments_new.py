# ==============================================
# 🚢 shipments_new.py - Novo registro de vendas
# ==============================================
 
# ---------- 1. Importações ----------
import streamlit as st
from database import load_df_udc, add_sales_record
from datetime import datetime, timedelta
import uuid
import time
import pandas as pd # Added for Excel upload
import unicodedata
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ---------- 2. Carregamento de dados externos ----------
df_udc = load_df_udc()
ports_pol_options = df_udc[df_udc["grupo"] == "Porto Origem"]["dado"].dropna().unique().tolist()
ports_pod_options = df_udc[df_udc["grupo"] == "Porto Destino"]["dado"].dropna().unique().tolist()
carrier_options = df_udc[df_udc["grupo"] == "Carrier"]["dado"].dropna().unique().tolist()
dthc_options = df_udc[df_udc["grupo"] == "DTHC"]["dado"].dropna().unique().tolist()
vip_pnl_risk_options = df_udc[df_udc["grupo"] == "VIP PNL Risk"]["dado"].dropna().unique().tolist()
# Carregar terminais da tabela F_ELLOX_TERMINALS (com fallback para unificada)
from database import list_terminal_names, list_terminal_names_from_unified
terminal_options = list_terminal_names() or list_terminal_names_from_unified() or []
 
# ---------- 3. Constantes ----------
# Mapeamento de colunas do Excel para campos internos do sistema
EXCEL_COLUMN_MAPPING = {
    "Referencia": "s_sales_order_reference",
    "Carrier": "b_voyage_carrier",
    "Origem": "s_port_of_loading_pol",
    "Destino_City": "s_port_of_delivery_pod",
    "Destino_Country": "b_pod_country_acronym",
    "Margem": "b_margin",
    "Afloat": "s_afloat",
    "Ctns": "s_quantity_of_containers",
    "Week": "s_requested_shipment_week",
    "Booking": "b_booking_reference",
    "Total Price": "b_freight_rate_usd",
    "Bogey": "b_bogey_sale_price_usd",
    "PnL Frete": "b_freightppnl",
    "PnL Bogey": "b_bogey_pnl",
    "ML": "b_ml_profit_margin",
    "Navio": "b_vessel_name",
    "Viagem": "b_voyage_code",
    "ETD": "b_data_estimativa_saida_etd",
    "Dthc": "s_dthc_prepaid",
    "Region": "b_destination_trade_region",
    "Ref_Sharepoint": "b_ref_sharepoint",
}

# Colunas obrigatórias do novo template
REQUIRED_EXCEL_COLS = ["Ref_Sharepoint", "Ctns", "Week", "Dthc"]

# Mapeamento de colunas do Excel para nomes de exibição padrão
EXCEL_DISPLAY_NAMES = {
    "Referencia": "SharePoint Reference",
    "Carrier": "Carrier",
    "Origem": "Port of Loading POL",
    "Destino_City": "Port of Delivery POD",
    "Destino_Country": "POD Country Acronym",
    "Margem": "Margin",
    "Afloat": "Afloat",
    "Ctns": "Quantity of Containers",
    "Week": "Requested Shipment Week",
    "Booking": "Booking Reference",
    "Total Price": "Freight Rate USD",
    "Bogey": "Bogey Sale Price USD",
    "PnL Frete": "Freight PNL",
    "PnL Bogey": "Bogey PNL",
    "ML": "ML Profit Margin",
    "Navio": "Vessel Name",
    "Viagem": "Voyage Code",
    "ETD": "ETD",
    "Dthc": "DTHC",
    "Region": "Destination Trade Region",
    "Ref_Sharepoint": "SharePoint Reference",
}

# Campos do formulário e seus nomes internos
form_fields = {
    "Type of Shipment": "s_type_of_shipment",
    "Quantity of Containers": "s_quantity_of_containers",
    "Port of Loading POL": "s_port_of_loading_pol",
    "Port of Delivery POD": "s_port_of_delivery_pod",
    "Final Destination": "s_final_destination",
    "Requested Shipment Week": "s_requested_shipment_week",
            "Requested Deadline Start Date": "s_requested_deadlines_start_date",
        "Requested Deadline End Date": "s_requested_deadlines_end_date",
    "Shipment Period Start Date": "s_shipment_period_start_date",
    "Shipment Period End Date": "s_shipment_period_end_date",
    "DTHC Prepaid": "s_dthc_prepaid",
    "Afloat": "s_afloat",
            "Required Arrival Date Expected": "s_required_arrival_date_expected",
    "Comments Sales": "s_comments"
}
 
# Campos obrigatórios
required_fields = {
    "s_type_of_shipment": "**:green[Type of Shipment]***",
    "s_quantity_of_containers": "**:green[Quantity of Containers]***",
    "s_port_of_delivery_pod": "**:green[Port of Delivery POD]***",
    "s_requested_shipment_week": "**:green[Requested Shipment Week]***",
    "s_dthc_prepaid": "**:green[DTHC Prepaid]***",
    "s_afloat": "**:green[Afloat]***"
}

# ---------- Funções de Normalização e Matching Inteligente ----------

# Mapeamento de abreviações conhecidas para carriers
CARRIER_ABBREVIATIONS = {
    "CMA": "CMA CGM",
    "CGM": "CMA CGM",
    "CMA-CGM": "CMA CGM",
    "HAPAG": "HAPAG-LLOYD",
    "LLOYD": "HAPAG-LLOYD",
    "HAPAG LLOYD": "HAPAG-LLOYD",
    "HLAG": "HAPAG-LLOYD",
}

def normalize_text_for_matching(text):
    """
    Normaliza texto para comparação removendo parênteses, acentos e normalizando espaços.
    
    Args:
        text: Texto a ser normalizado
    
    Returns:
        Texto normalizado em UPPERCASE
    """
    if not text or pd.isna(text):
        return ""
    
    text_str = str(text).strip()
    
    # Remove conteúdo entre parênteses
    text_str = re.sub(r'\s*\([^)]*\)', '', text_str)
    
    # Remove acentos (normaliza unicode e remove marcas diacríticas)
    text_str = unicodedata.normalize('NFD', text_str)
    text_str = ''.join(char for char in text_str if unicodedata.category(char) != 'Mn')
    
    # Normaliza espaços extras
    text_str = re.sub(r'\s+', ' ', text_str).strip()
    
    # Converte para UPPERCASE
    return text_str.upper()

def find_best_match(value, valid_options, field_type):
    """
    Encontra a melhor correspondência para um valor na lista de opções válidas.
    Usa estratégia de busca em múltiplas etapas para lidar com variações de formatação.
    
    Args:
        value: Valor a ser encontrado
        valid_options: Lista de opções válidas da base UDC
        field_type: Tipo do campo ("Port of Loading POL", "Port of Delivery POD", "Carrier")
    
    Returns:
        tuple: (valor_corrigido, is_valid, error_message)
    """
    if not value or pd.isna(value) or str(value).strip() == "":
        return "", True, None
    
    value_str = str(value).strip()
    value_normalized = normalize_text_for_matching(value_str)
    
    # Etapa 1: Busca exata case-insensitive (valor original vs. base)
    for option in valid_options:
        option_str = str(option).strip()
        if option_str.lower() == value_str.lower():
            return option_str, True, None  # Retorna valor exato da base
    
    # Etapa 2: Busca normalizada exata (após remover parênteses, acentos, etc.)
    for option in valid_options:
        option_str = str(option).strip()
        option_normalized = normalize_text_for_matching(option_str)
        if option_normalized == value_normalized and value_normalized != "":
            return option_str, True, None  # Retorna valor exato da base
    
    # Etapa 3: Busca parcial (valor contém ou está contido na opção)
    value_lower = value_str.lower()
    value_normalized_lower = value_normalized.lower()
    
    for option in valid_options:
        option_str = str(option).strip()
        option_lower = option_str.lower()
        option_normalized = normalize_text_for_matching(option_str)
        option_normalized_lower = option_normalized.lower()
        
        # Verifica se um contém o outro (case-insensitive)
        if (value_lower in option_lower or option_lower in value_lower) and len(value_lower) >= 3:
            return option_str, True, None
        
        # Verifica se um contém o outro (normalizado)
        if (value_normalized_lower in option_normalized_lower or option_normalized_lower in value_normalized_lower) and len(value_normalized_lower) >= 3:
            return option_str, True, None
    
    # Etapa 4: Busca por primeira palavra (para portos com nomes compostos)
    if field_type in ["Port of Loading POL", "Port of Delivery POD"]:
        value_first_word = value_normalized.split()[0] if value_normalized else ""
        if len(value_first_word) >= 3:
            for option in valid_options:
                option_str = str(option).strip()
                option_normalized = normalize_text_for_matching(option_str)
                option_first_word = option_normalized.split()[0] if option_normalized else ""
                if option_first_word == value_first_word:
                    return option_str, True, None
    
    # Etapa 5: Mapeamento de abreviações conhecidas (apenas para carriers)
    if field_type == "Carrier":
        value_upper = value_str.upper()
        if value_upper in CARRIER_ABBREVIATIONS:
            mapped_value = CARRIER_ABBREVIATIONS[value_upper]
            # Verifica se o valor mapeado existe na lista de opções válidas
            for option in valid_options:
                option_str = str(option).strip()
                if option_str.upper() == mapped_value.upper():
                    return option_str, True, None
    
    # Não encontrado - retorna valor normalizado (sem parênteses, em UPPERCASE) mas marca como inválido
    value_final = value_str.upper()  # Mantém formato UPPERCASE para consistência
    return value_final, False, f"{field_type} '{value_str}' não encontrado na base de dados"

# ---------- Função de Validação de Portos ----------
def validate_port_value(value, valid_options, port_type):
    """
    Valida se o valor do porto/carrier existe na lista de opções válidas.
    Usa matching inteligente para lidar com variações de formatação.
    
    Args:
        value: Valor do porto/carrier do Excel
        valid_options: Lista de opções válidas da base UDC
        port_type: Tipo do campo ("Port of Loading POL", "Port of Delivery POD", "Carrier")
    
    Returns:
        tuple: (valor_corrigido, is_valid, error_message)
    """
    return find_best_match(value, valid_options, port_type)

# ---------- Função para Detectar e Remover Linha de Cabeçalho ----------
def detect_and_remove_header_row(df):
    """
    Detecta se a primeira linha do DataFrame parece ser um cabeçalho/linha de ajuda
    e a remove se necessário.
    
    Critérios para detectar linha de cabeçalho:
    - Contém texto "Alt:" (indicando nomes alternativos)
    - Contém valores que parecem nomes de colunas de exibição
    - Muitos valores são strings que correspondem a EXCEL_DISPLAY_NAMES
    
    Args:
        df: DataFrame do Excel
    
    Returns:
        DataFrame sem a linha de cabeçalho (se detectada)
    """
    if df.empty or len(df) == 0:
        return df
    
    # Verificar primeira linha
    first_row = df.iloc[0]
    
    # Contar quantos valores da primeira linha parecem ser nomes de colunas
    display_names_values = list(EXCEL_DISPLAY_NAMES.values())
    matches_count = 0
    contains_alt = False
    
    for value in first_row.values:
        if pd.notna(value):
            value_str = str(value).strip()
            
            # Verificar se contém "Alt:"
            if "Alt:" in value_str or "alt:" in value_str.lower():
                contains_alt = True
            
            # Verificar se é um nome de exibição
            if value_str in display_names_values:
                matches_count += 1
    
    # Se mais de 50% dos valores correspondem ou contém "Alt:", é provavelmente linha de cabeçalho
    total_values = len([v for v in first_row.values if pd.notna(v)])
    if total_values > 0:
        match_ratio = matches_count / total_values
        
        # Se tem "Alt:" ou mais de 50% dos valores são nomes de exibição, remove
        if contains_alt or match_ratio > 0.5:
            return df.iloc[1:].reset_index(drop=True)
    
    return df

# ---------- Função para Gerar Template Excel ----------
def generate_excel_template():
    """
    Gera um arquivo Excel formatado com o template para upload em massa.
    
    Returns:
        bytes: Arquivo Excel em formato binário
    """
    # Ordem das colunas conforme template do usuário
    column_order = [
        "Ref_Sharepoint",
        "Carrier",
        "Origem",
        "Destino_City",
        "Destino_Country",
        "Margem",
        "Afloat",
        "Ctns",
        "Week",
        "Booking",
        "Total Price",
        "Bogey",
        "PnL Frete",
        "PnL Bogey",
        "ML",
        "Navio",
        "Viagem",
        "ETD",
        "Dthc",
        "Region",
    ]
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Estilos
    # Cabeçalho obrigatório (amarelo claro)
    header_required_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    # Cabeçalho opcional (azul claro)
    header_optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    # Texto de ajuda (cinza claro)
    help_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    bold_font = Font(bold=True, size=11)
    italic_font = Font(italic=True, size=9)
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Linha 1: Cabeçalhos (nomes originais para compatibilidade)
    for col_idx, col_name in enumerate(column_order, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Aplicar cor baseado se é obrigatória ou não
        if col_name in REQUIRED_EXCEL_COLS:
            cell.fill = header_required_fill
        else:
            cell.fill = header_optional_fill
    
    # Linha 2: Nomes alternativos (EXCEL_DISPLAY_NAMES)
    for col_idx, col_name in enumerate(column_order, start=1):
        display_name = EXCEL_DISPLAY_NAMES.get(col_name, col_name)
        cell = ws.cell(row=2, column=col_idx, value=f"Alt: {display_name}")
        cell.font = italic_font
        cell.alignment = center_alignment
        cell.fill = help_fill
        cell.border = thin_border
    
    # Linha 3: Linha de exemplo vazia
    for col_idx in range(1, len(column_order) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.border = thin_border
        cell.alignment = left_alignment
    
    # Ajustar largura das colunas
    for col_idx, col_name in enumerate(column_order, start=1):
        col_letter = get_column_letter(col_idx)
        # Ajustar largura baseado no nome da coluna
        max_length = max(len(col_name), len(EXCEL_DISPLAY_NAMES.get(col_name, "")) + 4)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    # Salvar em BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
 
# ---------- 4. Função principal ----------
def show_add_form():
    """
    Displays the form to add a new sales record.
    """
    st.subheader("New Sales Record 🚢")

    tab_manual, tab_excel = st.tabs(["Manual Entry", "Excel Upload (Bulk)"])

    with tab_manual:
        # ---------- Manual Form ----------
        if "current_farol_reference" not in st.session_state:
            st.session_state.current_farol_reference = str(uuid.uuid4())
            st.session_state.button_disabled = False

        if "confirm_disabled_until" not in st.session_state:
            st.session_state.confirm_disabled_until = None

        now = datetime.now()
        confirm_disabled = (
            st.session_state.confirm_disabled_until is not None and
            now < st.session_state.confirm_disabled_until
        )

        with st.form("add_form"):
            values = {}
            missing_fields = []

            col1, col2 = st.columns(2)
            with col1:
                values["s_farol_status"] = st.selectbox("**:green[Farol Status]***", ["New request"], index=0, disabled=True)

            col1, col2 = st.columns(2)
            with col1:
                values["s_type_of_shipment"] = st.selectbox("**:green[Type of Shipment]***", ["", "Forecast", "Extra"])
            with col2:
                values["s_quantity_of_containers"] = st.number_input("**:green[Quantity of Containers]***", min_value=0, step=1)

            # Linha 1: Requested Shipment Week com mesma largura dos demais (1/2 da linha)
            col1, col2 = st.columns(2)
            with col1:
                values["s_requested_shipment_week"] = st.number_input("**:green[Requested Shipment Week]***", min_value=1, max_value=52, step=1)
            with col2:
                st.empty()

            # Linha 2: Required Sail Date | Required Arrival Date
            col1, col2 = st.columns(2)
            with col1:
                values["s_required_sail_date"] = st.date_input(
                    "Required Sail Date",
                    value=None,
                    key=f"add_required_sail_date_{st.session_state.current_farol_reference}"
                )
            with col2:
                values["s_required_arrival_date_expected"] = st.date_input(
                    "Required Arrival Date",
                    value=None,
                    key=f"add_required_arrival_date_{st.session_state.current_farol_reference}"
                )

            # Linha 3: Requested Deadline Start | Requested Deadline End
            col1, col2 = st.columns(2)
            with col1:
                values["s_requested_deadlines_start_date"] = st.date_input(
                    "Requested Deadline Start Date",
                    value=None,
                    key=f"add_requested_deadline_start_{st.session_state.current_farol_reference}"
                )
            with col2:
                values["s_requested_deadlines_end_date"] = st.date_input(
                    "Requested Deadline End Date",
                    value=None,
                    key=f"add_requested_deadline_end_{st.session_state.current_farol_reference}"
                )

            # Linha 4: Shipment Period Start | Shipment Period End
            col1, col2 = st.columns(2)
            with col1:
                values["s_shipment_period_start_date"] = st.date_input(
                    "Shipment Period Start Date",
                    value=None,
                    key=f"add_shipment_period_start_{st.session_state.current_farol_reference}"
                )
            with col2:
                values["s_shipment_period_end_date"] = st.date_input(
                    "Shipment Period End Date",
                    value=None,
                    key=f"add_shipment_period_end_{st.session_state.current_farol_reference}"
                )

            col1, col2 = st.columns(2)
            with col1:
                values["s_shipment_period_start_date"] = st.date_input("Shipment Period Start Date", value=None)
            with col2:
                values["s_shipment_period_end_date"] = st.date_input("Shipment Period End Date", value=None)

            # (Mantido sem campo adicional aqui; já foi exibido logo após Required Sail Date)

            col1, col2 = st.columns(2)
            with col1:
                values["s_dthc_prepaid"] = st.selectbox("**:green[DTHC]***", [""] + dthc_options)
            with col2:
                values["s_afloat"] = st.selectbox("**:green[Afloat]***", ["", "Yes", "No"])

            col1, col2 = st.columns(2)
            with col1:
                values["s_port_of_loading_pol"] = st.selectbox("Port of Loading POL", [""] + ports_pol_options)
            with col2:
                values["s_port_of_delivery_pod"] = st.selectbox("**:green[Port of Delivery POD]***", [""] + ports_pod_options)

            col1, col2 = st.columns(2)
            with col1:
                values["s_final_destination"] = st.text_input("Final Destination")
            with col2:
                values["b_terminal"] = st.selectbox("Terminal", [""] + terminal_options)
            
            # Restriction Type na linha de baixo, ocupando 1 coluna
            col1, col2 = st.columns(2)
            with col1:
                values["s_vip_pnl_risk"] = st.selectbox("Restriction Type", [""] + vip_pnl_risk_options)

            values["s_comments"] = st.text_area("Sales Comments")

            for field, label in required_fields.items():
                if not values.get(field):
                    missing_fields.append(label)

            col1, col2 = st.columns(2)
            with col1:
                salvar = st.form_submit_button(
                    "✅ Confirm",
                    disabled=st.session_state.get("button_disabled", False)
                )
            with col2:
                voltar = st.form_submit_button("🔙 Back to Shipments")

            if salvar and not st.session_state.get("button_disabled", False):
                if missing_fields:
                    st.error(f"Please fill in the required fields: {', '.join(missing_fields)}")
                else:
                    st.session_state.button_disabled = True
                    values["adjustment_id"] = st.session_state.current_farol_reference
                    values["user_insert"] = ''
                    
                    # Mapeamento do s_pnl_destination baseado no s_vip_pnl_risk
                    values["s_pnl_destination"] = "Yes" if values.get("s_vip_pnl_risk") == "PNL" else "No"
                    
                    # REMOVIDO: Mapeamento incorreto que colocava VIP/PNL/RISK em s_final_destination
                    # O campo s_final_destination deve ser preenchido manualmente pelo usuário
                    # O campo s_vip_pnl_risk é independente e não deve afetar s_final_destination
                    
                    # Tratamento da data s_required_arrival_date_expected no formulário manual
                    if values.get("s_required_arrival_date_expected"):
                        try:
                            # Converte para datetime se for uma data válida
                            if isinstance(values["s_required_arrival_date_expected"], str):
                                values["s_required_arrival_date_expected"] = pd.to_datetime(values["s_required_arrival_date_expected"])
                        except (ValueError, TypeError):
                            values["s_required_arrival_date_expected"] = None

                    with st.spinner("Processing new shipment, please wait..."):
                        if add_sales_record(values):
                            st.success("✅ Data saved successfully!")
                            time.sleep(2)
                            st.session_state.pop("current_farol_reference", None)
                            st.session_state.pop("button_disabled", None)
                            st.session_state["current_page"] = "main"
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.session_state.button_disabled = False
                            st.error("Error saving data. Please try again.")

            elif voltar:
                st.session_state.pop("current_farol_reference", None)
                st.session_state.pop("button_disabled", None)
                st.session_state["current_page"] = "main"
                st.rerun()

    with tab_excel:
        # Botão de download do template
        template_data = generate_excel_template()
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Download the Excel template to ensure the correct format:**")
        with col2:
            st.download_button(
                label="📥 Download Template",
                data=template_data,
                file_name=f"template_forecast_upload_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_template_btn"
            )
        
        st.markdown("---")
        uploaded_file = st.file_uploader("Select an Excel file (.xlsx)", type=["xlsx"], key="excel_mass_upload")
        
        # Manter df_excel no session_state para evitar perda de estado ao recarregar
        if "df_excel_upload" not in st.session_state:
            st.session_state.df_excel_upload = None
        
        df_excel = st.session_state.df_excel_upload
        
        # Processa e exibe o arquivo automaticamente quando carregado
        if uploaded_file:
            try:
                df_excel = pd.read_excel(uploaded_file)
                
                # Detectar e remover linha de cabeçalho (linha 0) se presente
                df_excel = detect_and_remove_header_row(df_excel)
                
                # Salvar no session_state para manter estado
                st.session_state.df_excel_upload = df_excel
                
                # Validar portos e carrier durante carregamento para destacar células inválidas
                invalid_port_cells = []  # Lista de células inválidas: (row_idx, col_name)
                
                if "Origem" in df_excel.columns:
                    for idx, row in df_excel.iterrows():
                        port_value = row.get("Origem", "")
                        if pd.notna(port_value) and str(port_value).strip() != "":
                            _, is_valid, _ = validate_port_value(port_value, ports_pol_options, "Port of Loading POL")
                            if not is_valid:
                                invalid_port_cells.append((idx, "Origem"))
                
                if "Destino_City" in df_excel.columns:
                    for idx, row in df_excel.iterrows():
                        port_value = row.get("Destino_City", "")
                        if pd.notna(port_value) and str(port_value).strip() != "":
                            _, is_valid, _ = validate_port_value(port_value, ports_pod_options, "Port of Delivery POD")
                            if not is_valid:
                                invalid_port_cells.append((idx, "Destino_City"))
                
                if "Carrier" in df_excel.columns:
                    for idx, row in df_excel.iterrows():
                        carrier_value = row.get("Carrier", "")
                        if pd.notna(carrier_value) and str(carrier_value).strip() != "":
                            _, is_valid, _ = validate_port_value(carrier_value, carrier_options, "Carrier")
                            if not is_valid:
                                invalid_port_cells.append((idx, "Carrier"))
                
                # Criar cópia para exibição com nomes padrão
                df_display = df_excel.copy()
                
                # Renomear colunas para nomes de exibição padrão
                rename_dict = {}
                for col in df_display.columns:
                    if col in EXCEL_DISPLAY_NAMES:
                        rename_dict[col] = EXCEL_DISPLAY_NAMES[col]
                
                if rename_dict:
                    df_display.rename(columns=rename_dict, inplace=True)
                
                # Garantir que não há colunas duplicadas (pode acontecer se múltiplas colunas mapearem para o mesmo nome)
                if df_display.columns.duplicated().any():
                    # Remover colunas duplicadas, mantendo apenas a primeira ocorrência
                    df_display = df_display.loc[:, ~df_display.columns.duplicated()]
                
                # Garantir que o índice seja único
                df_display = df_display.reset_index(drop=True)
                
                # Função para destacar colunas importantes e células inválidas
                def highlight_cols_and_invalid_ports(row):
                    styles = []
                    for col_name in df_display.columns:
                        # Verificar se é coluna importante
                        is_important = col_name in list(EXCEL_DISPLAY_NAMES.values())
                        
                        # Verificar se é célula de porto inválida
                        is_invalid_port = False
                        row_idx = row.name
                        
                        # Verificar se esta célula é um porto ou carrier inválido
                        for invalid_row_idx, invalid_col in invalid_port_cells:
                            if invalid_row_idx == row_idx:
                                # Mapear coluna original para coluna de exibição
                                if invalid_col == "Origem" and col_name == "Port of Loading POL":
                                    is_invalid_port = True
                                    break
                                elif invalid_col == "Destino_City" and col_name == "Port of Delivery POD":
                                    is_invalid_port = True
                                    break
                                elif invalid_col == "Carrier" and col_name == "Carrier":
                                    is_invalid_port = True
                                    break
                        
                        # Aplicar estilo
                        if is_invalid_port:
                            styles.append('background-color: #ffcccc; color: #cc0000; font-weight: bold;')
                        elif is_important:
                            styles.append('background-color: #e6f3ff; font-weight: bold;')
                        else:
                            styles.append('')
                    return styles
                
                # CSS para remover padding das células do DataFrame
                st.markdown("""
                <style>
                div[data-testid="stDataFrame"] table td,
                div[data-testid="stDataFrame"] table th {
                    padding: 0 !important;
                }
                div[data-testid="stDataFrame"] table {
                    border-collapse: collapse;
                }
                </style>
                """, unsafe_allow_html=True)
                
                st.dataframe(df_display.style.apply(highlight_cols_and_invalid_ports, axis=1))
                
                # Exibir aviso se houver portos ou carriers inválidos
                if invalid_port_cells:
                    invalid_count = len(invalid_port_cells)
                    invalid_items = []
                    for _, col in invalid_port_cells:
                        if col == "Origem":
                            invalid_items.append("Port of Loading POL")
                        elif col == "Destino_City":
                            invalid_items.append("Port of Delivery POD")
                        elif col == "Carrier":
                            invalid_items.append("Carrier")
                    items_text = ", ".join(set(invalid_items))
                    st.warning(f"⚠️ {invalid_count} valor(es) não encontrado(s) na base de dados (destacados em vermelho): {items_text}. Corrija os valores antes de confirmar o upload.")
                
                # Validação das colunas obrigatórias
                missing_cols = [col for col in REQUIRED_EXCEL_COLS if col not in df_excel.columns]
                if missing_cols:
                    st.error(f"The file is missing the required columns: {', '.join(missing_cols)}")
                else:
                    st.success(f"✅ File loaded successfully! Found {len(df_excel)} rows to process.")
                    
            except Exception as e:
                st.error(f"Error processing file: {str(e)}")
        
        # Formulário apenas para os botões
        with st.form("excel_upload_box"):
            col1, col2 = st.columns(2)
            with col1:
                confirm_bulk = st.form_submit_button("✅ Confirm Bulk Upload", disabled=(df_excel is None))
            with col2:
                back_bulk = st.form_submit_button("🔙 Back to Shipments")
            
            if back_bulk:
                st.session_state.pop("df_excel_upload", None)
                st.session_state["current_page"] = "main"
                st.rerun()
            
            if confirm_bulk and df_excel is not None:
                success, fail = 0, 0
                port_validation_errors = []  # Coletar erros de validação de portos
                progress_bar = st.progress(0, text="Processing shipments...")
                
                for idx, row in df_excel.iterrows():
                    values = {
                        # Campos padrão
                        "s_farol_status": "New request",
                        "s_type_of_shipment": "Forecast",
                        "adjustment_id": str(uuid.uuid4()),
                        "user_insert": '',
                    }
                    
                    # Mapear todas as colunas do Excel para os campos internos
                    for excel_col, internal_field in EXCEL_COLUMN_MAPPING.items():
                        if excel_col in df_excel.columns:
                            cell_value = row.get(excel_col, "")
                            
                            # Tratamento especial para diferentes tipos de dados
                            if internal_field in ["s_quantity_of_containers"]:
                                # Converter para inteiro
                                try:
                                    values[internal_field] = int(float(cell_value)) if pd.notna(cell_value) and str(cell_value).strip() != "" else 0
                                except (ValueError, TypeError):
                                    values[internal_field] = 0
                            
                            elif internal_field in ["b_data_estimativa_saida_etd"]:
                                # Converter datas
                                if pd.notna(cell_value) and str(cell_value).strip() != "":
                                    try:
                                        if isinstance(cell_value, str):
                                            values[internal_field] = pd.to_datetime(cell_value)
                                        elif isinstance(cell_value, (int, float)):
                                            values[internal_field] = None
                                        else:
                                            values[internal_field] = cell_value
                                    except (ValueError, TypeError):
                                        values[internal_field] = None
                                else:
                                    values[internal_field] = None
                            
                            elif internal_field in ["b_margin", "b_freight_rate_usd", "b_bogey_sale_price_usd", 
                                                     "b_freightppnl", "b_bogey_pnl", "b_ml_profit_margin"]:
                                # Converter valores numéricos monetários
                                try:
                                    if pd.notna(cell_value) and str(cell_value).strip() != "":
                                        values[internal_field] = float(cell_value)
                                    else:
                                        values[internal_field] = None
                                except (ValueError, TypeError):
                                    values[internal_field] = None
                            
                            else:
                                # Campos de texto simples
                                if pd.notna(cell_value):
                                    values[internal_field] = str(cell_value).strip()
                                else:
                                    values[internal_field] = ""
                    
                    # Validação de portos após mapeamento
                    # Validar Port of Loading POL (Origem)
                    if "s_port_of_loading_pol" in values:
                        original_value = values["s_port_of_loading_pol"]
                        corrected_value, is_valid, error_msg = validate_port_value(
                            original_value,
                            ports_pol_options,
                            "Port of Loading POL"
                        )
                        values["s_port_of_loading_pol"] = corrected_value
                        if not is_valid:
                            port_validation_errors.append({
                                "row": idx + 1,
                                "column": "Origem",
                                "value": original_value,
                                "corrected_value": corrected_value,
                                "message": error_msg
                            })
                    
                    # Validar Port of Delivery POD (Destino_City)
                    if "s_port_of_delivery_pod" in values:
                        original_value = values["s_port_of_delivery_pod"]
                        corrected_value, is_valid, error_msg = validate_port_value(
                            original_value,
                            ports_pod_options,
                            "Port of Delivery POD"
                        )
                        values["s_port_of_delivery_pod"] = corrected_value
                        if not is_valid:
                            port_validation_errors.append({
                                "row": idx + 1,
                                "column": "Destino_City",
                                "value": original_value,
                                "corrected_value": corrected_value,
                                "message": error_msg
                            })
                    
                    # Validar Carrier
                    if "b_voyage_carrier" in values:
                        original_value = values["b_voyage_carrier"]
                        corrected_value, is_valid, error_msg = validate_port_value(
                            original_value,
                            carrier_options,
                            "Carrier"
                        )
                        values["b_voyage_carrier"] = corrected_value
                        if not is_valid:
                            port_validation_errors.append({
                                "row": idx + 1,
                                "column": "Carrier",
                                "value": original_value,
                                "corrected_value": corrected_value,
                                "message": error_msg
                            })
                    
                    # Validação de campos obrigatórios
                    required_check_fields = {
                        "b_ref_sharepoint": "Ref_Sharepoint",
                        "s_quantity_of_containers": "Ctns",
                        "s_requested_shipment_week": "Week",
                        "s_dthc_prepaid": "Dthc",
                    }
                    
                    missing_required = []
                    for field, excel_col in required_check_fields.items():
                        if not values.get(field) or (isinstance(values.get(field), (int, float)) and values.get(field) == 0):
                            missing_required.append(excel_col)
                    
                    if missing_required:
                        fail += 1
                        continue
                    
                    try:
                        if add_sales_record(values):
                            success += 1
                        else:
                            fail += 1
                    except Exception as e:
                        fail += 1
                        print(f"Erro ao processar linha {idx + 1}: {e}")
                    
                    # Atualiza a barra de progresso
                    progress = (idx + 1) / len(df_excel)
                    progress_bar.progress(progress, text=f"Processing shipment {idx+1} of {len(df_excel)}...")
                
                progress_bar.empty()
                
                # Exibir warnings de validação de portos e carriers se houver
                if port_validation_errors:
                    error_types = {}
                    for error in port_validation_errors:
                        col_type = error['column']
                        if col_type not in error_types:
                            error_types[col_type] = []
                        error_types[col_type].append(error)
                    
                    error_summary = []
                    for col_type, errors in error_types.items():
                        error_summary.append(f"{len(errors)} {col_type}(s)")
                    
                    st.warning(f"⚠️ {len(port_validation_errors)} valor(es) não encontrado(s) na base de dados ({', '.join(error_summary)}):")
                    error_details = []
                    for error in port_validation_errors:
                        error_details.append(f"Linha {error['row']}, Coluna '{error['column']}': Valor '{error['value']}' → Corrigido para '{error['corrected_value']}' (não encontrado na base)")
                    st.text("\n".join(error_details))
                    st.info("💡 Os valores foram capitalizados automaticamente. Verifique se os valores estão corretos na base UDC ou se precisam ser adicionados.")
                
                st.success(f"✅ {success} shipments successfully uploaded!")
                if fail:
                    st.error(f"❌ {fail} shipments failed. Please check the file data.")
                
                # Limpar estado do upload após processamento
                st.session_state.pop("df_excel_upload", None)
                
                # Limpa o cache e volta para a tela principal
                st.cache_data.clear()
                time.sleep(2)
                st.session_state["current_page"] = "main"
                st.rerun()